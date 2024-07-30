import tkinter as tk
from tkinter import filedialog, scrolledtext
from tkinter import ttk
from tkhtmlview import HTMLLabel
from openpyxl import load_workbook
from datetime import datetime
from tqdm import tqdm
import threading

def calculate_average_temperature(file_path, start_row, progress_var):
    wb = load_workbook(filename=file_path)
    sheet = wb.active   
    date_temperature_sum = {}
    date_temperature_count = {}
    
    total_rows = sheet.max_row - start_row + 1
    row_count = 0

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        date = datetime.strptime(row[0], "%d.%m.%Y %H:%M").date()
        temperature = row[1]       
        if temperature is not None:
            if date not in date_temperature_sum:
                date_temperature_sum[date] = 0
                date_temperature_count[date] = 0            
            date_temperature_sum[date] += temperature
            date_temperature_count[date] += 1
        
        row_count += 1
        progress_var.set((row_count / total_rows) * 100)
    
    average_temperature_per_date = {date: date_temperature_sum[date] / date_temperature_count[date] 
                                    for date in date_temperature_sum}    
    return average_temperature_per_date

def count_days_above_threshold(average_temperature_per_date, threshold, progress_var):
    days_above_threshold = 0
    total_dates = len(average_temperature_per_date)
    date_count = 0

    for date, temp in average_temperature_per_date.items():
        if temp >= threshold:
            days_above_threshold += 1
        date_count += 1
        progress_var.set((date_count / total_dates) * 100)
    
    return days_above_threshold

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_label.config(text=file_path)
        file_label.file_path = file_path

def run_calculation():
    file_path = getattr(file_label, 'file_path', None)
    if not file_path:
        result_label.set_html("<h2>Выберите файл!</h2>")
        return
    
    try:
        start_row = int(start_row_entry.get())
        threshold = float(threshold_entry.get())
    except ValueError:
        result_label.set_html("<h2>Некорректные входные данные!</h2>")
        return

    progress_var.set(0)
    progress_bar.start()

    def task():
        average_temperature_per_date = calculate_average_temperature(file_path, start_row, progress_var)
        days_above_threshold = count_days_above_threshold(average_temperature_per_date, threshold, progress_var)

        date_text.delete(1.0, tk.END)
        threshold_text.delete(1.0, tk.END)
        
        for date, temp in average_temperature_per_date.items():
            date_text.insert(tk.END, f"{date}: {temp:.2f}\n")
        
        threshold_text.insert(tk.END, f"Количество дней с температурой выше {threshold} градусов: {days_above_threshold}")
        progress_bar.stop()

    threading.Thread(target=task).start()

root = tk.Tk()
root.title("Расчет дней с определенной T по базам XLSX")
root.geometry("650x900")

style = ttk.Style(root)
style.theme_use('clam')

frame = ttk.Frame(root, padding="5")
frame.pack(fill=tk.X)

file_label = ttk.Label(frame, text="Выберите файл", width=50)
file_label.grid(row=0, column=0, columnspan=2, pady=3, sticky="ew")

browse_button = ttk.Button(frame, text="Обзор", command=browse_file)
browse_button.grid(row=0, column=2, pady=3, padx=3)

ttk.Label(frame, text="С какой строки начать:").grid(row=1, column=0, pady=3, sticky="w")
start_row_entry = ttk.Entry(frame)
start_row_entry.grid(row=1, column=1, pady=3, sticky="ew", padx=3)

ttk.Label(frame, text="Температура:").grid(row=2, column=0, pady=3, sticky="w")
threshold_entry = ttk.Entry(frame)
threshold_entry.grid(row=2, column=1, pady=3, sticky="ew", padx=3)

run_button = ttk.Button(frame, text="Запуск", command=run_calculation)
run_button.grid(row=3, column=0, columnspan=3, pady=5)

progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(frame, variable=progress_var, maximum=100)
progress_bar.grid(row=4, column=0, columnspan=3, pady=5, sticky="ew")

result_label = HTMLLabel(root, html="")
result_label.pack(padx=5, pady=5, fill=tk.BOTH, expand=True)

console_frame = ttk.Frame(root, padding="5")
console_frame.pack(fill=tk.BOTH, expand=True)

date_text = scrolledtext.ScrolledText(console_frame, width=50, height=10)
date_text.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

threshold_text = scrolledtext.ScrolledText(console_frame, width=50, height=10)
threshold_text.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")

console_frame.columnconfigure(0, weight=1)
console_frame.columnconfigure(1, weight=1)
console_frame.rowconfigure(0, weight=1)

root.mainloop()
